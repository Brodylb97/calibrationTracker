# ui/main_window.py - Main application window

from datetime import date, datetime
import csv
from pathlib import Path

from PyQt5 import QtWidgets, QtCore, QtGui

from database import CalibrationRepository, get_effective_db_path, get_connection, DB_PATH, persist_last_db_path, StaleDataError
from services import instrument_service
from lan_notify import send_due_reminders_via_lan

from ui.table_models import InstrumentTableModel, InstrumentFilterProxyModel, HighlightDelegate
from ui.theme import (
    get_saved_theme, set_saved_theme, get_all_themes, get_saved_font_size, set_saved_font_size,
    FONT_SIZE_OPTIONS, _app_icon_path, apply_global_style, get_theme_colors,
)
from ui.help_content import HelpDialog
from ui.dialogs import (
    InstrumentDialog, SettingsDialog, AttachmentsDialog,
    DestinationEditDialog, DestinationsDialog,
    PersonnelEditDialog, PersonnelDialog,
    TemplateEditDialog, FieldEditDialog, ExplainToleranceDialog,
    TemplateFieldsDialog, TemplatesDialog,
    CalibrationHistoryDialog, CalibrationFormDialog,
    BatchUpdateDialog, BatchAssignInstrumentTypeDialog,
    CalDateDialog, InstrumentInfoDialog, AuditLogDialog,
    ThemeEditorDialog,
)


class ExportWorker(QtCore.QThread):
    """Background worker for PDF export with progress and cancel support."""
    progress = QtCore.pyqtSignal(int, int)
    finished = QtCore.pyqtSignal(dict)
    error = QtCore.pyqtSignal(str)

    def __init__(self, db_path: Path, target_dir: str | Path):
        super().__init__()
        self.db_path = db_path
        self.target_dir = Path(target_dir)
        self._cancelled = False

    def cancel(self):
        self._cancelled = True

    def run(self):
        try:
            from database import get_connection, initialize_db, CalibrationRepository
            from pdf_export import export_all_calibrations_to_directory
            conn = get_connection(self.db_path)
            conn = initialize_db(conn, self.db_path)
            repo = CalibrationRepository(conn)
            try:
                def on_progress(cur, tot):
                    self.progress.emit(cur, tot)
                result = export_all_calibrations_to_directory(
                    repo, self.target_dir,
                    progress_callback=on_progress,
                    cancelled_check=lambda: self._cancelled,
                )
                self.finished.emit(result)
            finally:
                conn.close()
        except Exception as e:
            self.error.emit(str(e))


class MainWindow(QtWidgets.QMainWindow):
    def __init__(self, repo: CalibrationRepository):
        super().__init__()
        self.repo = repo
        self.setWindowTitle("Calibration Tracker")
        self.resize(1000, 600)
        icon_path = _app_icon_path()
        if icon_path.is_file():
            self.setWindowIcon(QtGui.QIcon(str(icon_path)))

        self._init_ui()
        self.load_instruments()

    def _init_ui(self):
        central = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(central)

        # ------------------------------------------------------------------
        # Toolbar - Streamlined with most common actions
        # ------------------------------------------------------------------
        toolbar = self.addToolBar("Main")
        toolbar.setMovable(False)
        toolbar.setToolButtonStyle(QtCore.Qt.ToolButtonTextBesideIcon)

        # Primary actions - most frequently used
        self.act_new = toolbar.addAction("New")
        self.act_new.setShortcut(QtGui.QKeySequence.New)
        self.act_new.setToolTip("Create a new instrument (Ctrl+N)")

        self.act_edit = toolbar.addAction("Edit")
        self.act_edit.setShortcut(QtGui.QKeySequence("Ctrl+E"))
        self.act_edit.setToolTip("Edit selected instrument (Ctrl+E)")

        toolbar.addSeparator()

        # Calibration actions
        self.act_cal = toolbar.addAction("Mark Calibrated")
        self.act_cal.setShortcut(QtGui.QKeySequence("Ctrl+M"))
        self.act_cal.setToolTip("Mark selected instrument as calibrated (Ctrl+M)")

        self.act_hist = toolbar.addAction("History")
        self.act_hist.setShortcut(QtGui.QKeySequence("Ctrl+H"))
        self.act_hist.setToolTip("View calibration history (Ctrl+H)")

        toolbar.addSeparator()

        # Settings (moved to end, less frequently used)
        self.act_settings = toolbar.addAction("Settings")
        self.act_settings.setShortcut(QtGui.QKeySequence.Preferences)
        self.act_settings.setToolTip("Open settings (Ctrl+,)")

        # Connect toolbar actions
        self.act_new.triggered.connect(self.on_new)
        self.act_edit.triggered.connect(self.on_edit)
        self.act_cal.triggered.connect(self.on_mark_calibrated)
        self.act_hist.triggered.connect(self.on_cal_history)
        self.act_settings.triggered.connect(self.on_settings)
        
        # Actions not in toolbar (available via menus/context menu)
        self.act_delete = QtWidgets.QAction("Delete", self)
        self.act_delete.setShortcut(QtGui.QKeySequence.Delete)
        self.act_delete.setToolTip("Delete selected instrument (Del)")
        self.act_delete.triggered.connect(self.on_delete)
        
        self.act_view_info = QtWidgets.QAction("View Details", self)
        self.act_view_info.setShortcut(QtGui.QKeySequence("Ctrl+I"))
        self.act_view_info.setToolTip("View detailed information (Ctrl+I)")
        self.act_view_info.triggered.connect(self.on_view_info)
        
        self.act_templates = QtWidgets.QAction("Templates", self)
        self.act_templates.setShortcut(QtGui.QKeySequence("Ctrl+T"))
        self.act_templates.setToolTip("Manage calibration templates (Ctrl+T)")
        self.act_templates.triggered.connect(self.on_templates)
        
        self.act_dest = QtWidgets.QAction("Destinations", self)
        self.act_dest.setToolTip("Manage calibration destinations")
        self.act_dest.triggered.connect(self.on_destinations)

        self.act_personnel = QtWidgets.QAction("Personnel", self)
        self.act_personnel.setToolTip("Manage personnel (technicians authorized to perform calibrations)")
        self.act_personnel.triggered.connect(self.on_personnel)
        
        self.act_reminders = QtWidgets.QAction("Send Reminders", self)
        self.act_reminders.setToolTip("Send LAN reminders for due calibrations")
        self.act_reminders.triggered.connect(self.on_send_reminders)

        self.act_batch_update = QtWidgets.QAction("Batch update...", self)
        self.act_batch_update.setToolTip("Update status or next due date for selected instruments")
        self.act_batch_update.triggered.connect(self.on_batch_update)
        self.act_batch_assign_type = QtWidgets.QAction("Batch assign instrument type...", self)
        self.act_batch_assign_type.setToolTip("Assign the same instrument type to selected instruments")
        self.act_batch_assign_type.triggered.connect(self.on_batch_assign_type)

        # ------------------------------------------------------------------
        # Menus (grouped by domain: File, Edit, Calibrations, Tools, View, Help)
        # ------------------------------------------------------------------
        menubar = self.menuBar()

        # File - create, export, exit (session/document level)
        file_menu = menubar.addMenu("&File")
        file_menu.addAction(self.act_new)
        file_menu.addSeparator()
        export_menu = file_menu.addMenu("&Export")
        export_csv_action = export_menu.addAction("Current view to CSV...")
        export_csv_action.setShortcut(QtGui.QKeySequence("Ctrl+Shift+C"))
        export_csv_action.triggered.connect(self.on_export_csv)
        export_excel_action = export_menu.addAction("Current view to Excel...")
        export_excel_action.setShortcut(QtGui.QKeySequence("Ctrl+Shift+E"))
        export_excel_action.triggered.connect(self.on_export_excel)
        export_menu.addSeparator()
        export_pdf_action = export_menu.addAction("All calibrations to PDF...")
        export_pdf_action.setShortcut(QtGui.QKeySequence("Ctrl+Shift+P"))
        export_pdf_action.triggered.connect(self.on_export_all_calibrations)
        export_menu.addSeparator()
        export_internal_action = export_menu.addAction("Export preset: Internal review...")
        export_internal_action.setToolTip("Export all calibrations to PDF in Internal_Review_<date> subfolder")
        export_internal_action.triggered.connect(lambda: self._on_export_preset("Internal_Review"))
        export_audit_action = export_menu.addAction("Export preset: External audit...")
        export_audit_action.setToolTip("Export all calibrations to PDF in External_Audit_<date> subfolder")
        export_audit_action.triggered.connect(lambda: self._on_export_preset("External_Audit"))
        export_customer_action = export_menu.addAction("Export preset: Customer delivery...")
        export_customer_action.setToolTip("Export all calibrations to PDF in Customer_<date> subfolder")
        export_customer_action.triggered.connect(lambda: self._on_export_preset("Customer"))
        file_menu.addSeparator()
        exit_action = file_menu.addAction("E&xit")
        exit_action.setShortcut(QtGui.QKeySequence.Quit)
        exit_action.triggered.connect(self.close)

        # Edit - selection actions (edit, view, batch, delete)
        edit_menu = menubar.addMenu("&Edit")
        edit_menu.addAction(self.act_edit)
        edit_menu.addAction(self.act_view_info)
        edit_menu.addSeparator()
        edit_menu.addAction(self.act_batch_update)
        edit_menu.addAction(self.act_batch_assign_type)
        edit_menu.addSeparator()
        edit_menu.addAction(self.act_delete)

        # Calibrations - calibration records and procedures
        cal_menu = menubar.addMenu("&Calibrations")
        cal_menu.addAction(self.act_cal)
        cal_menu.addAction(self.act_hist)
        cal_menu.addSeparator()
        cal_menu.addAction(self.act_templates)

        # Tools - reference data and operations
        tools_menu = menubar.addMenu("&Tools")
        tools_menu.addAction(self.act_dest)
        tools_menu.addAction(self.act_personnel)
        tools_menu.addAction(self.act_reminders)
        tools_menu.addSeparator()
        tools_menu.addAction(self.act_settings)

        # View - appearance (theme, text size)
        view_menu = menubar.addMenu("&View")
        self._theme_sub = view_menu.addMenu("&Theme")
        self._theme_action_group = QtWidgets.QActionGroup(self)
        self._theme_action_group.setExclusive(True)
        self._theme_sub.aboutToShow.connect(self._populate_theme_menu)
        self._populate_theme_menu()
        text_size_sub = view_menu.addMenu("&Text Size")
        self._text_size_action_group = QtWidgets.QActionGroup(self)
        self._text_size_action_group.setExclusive(True)
        current_pt = get_saved_font_size()
        for label, pt in FONT_SIZE_OPTIONS:
            act = text_size_sub.addAction(label)
            act.setCheckable(True)
            act.setChecked(pt == current_pt)
            act.setData(pt)
            act.triggered.connect(lambda checked, p=pt: self._on_text_size_selected(p))
            self._text_size_action_group.addAction(act)

        # Help - documentation, refresh, about
        help_menu = menubar.addMenu("&Help")
        shortcuts_action = help_menu.addAction("Keyboard Shortcuts...")
        shortcuts_action.setShortcut(QtGui.QKeySequence("F1"))
        shortcuts_action.triggered.connect(self.on_show_shortcuts)
        help_menu.addSeparator()
        refresh_db_action = help_menu.addAction("Refresh database (reconnect to server)")
        refresh_db_action.setToolTip("Reconnect to the server database (app uses server only, no local copy).")
        refresh_db_action.triggered.connect(self._on_refresh_database)
        help_menu.addSeparator()
        about_action = help_menu.addAction("About...")
        about_action.triggered.connect(self.on_show_about)

        # ------------------------------------------------------------------
        # Needs Attention panel (overdue, due soon, recently modified, last cal failed)
        # ------------------------------------------------------------------
        self._needs_attention_container = QtWidgets.QWidget()
        needs_layout = QtWidgets.QHBoxLayout(self._needs_attention_container)
        needs_layout.setContentsMargins(8, 4, 8, 4)
        needs_layout.setSpacing(8)
        needs_layout.addWidget(QtWidgets.QLabel("Needs Attention:"))
        self._btn_overdue = QtWidgets.QPushButton("Overdue (0)")
        self._btn_overdue.setCheckable(True)
        self._btn_overdue.setMinimumWidth(95)
        self._btn_overdue.setToolTip("Show instruments past due date")
        self._btn_overdue.clicked.connect(self._on_needs_attention_overdue)
        needs_layout.addWidget(self._btn_overdue)
        self._btn_due_soon = QtWidgets.QPushButton("Due soon (0)")
        self._btn_due_soon.setCheckable(True)
        self._btn_due_soon.setMinimumWidth(95)
        self._btn_due_soon.setToolTip("Show instruments due within 30 days")
        self._btn_due_soon.clicked.connect(self._on_needs_attention_due_soon)
        needs_layout.addWidget(self._btn_due_soon)
        self._btn_recently_modified = QtWidgets.QPushButton("Recently modified (0)")
        self._btn_recently_modified.setCheckable(True)
        self._btn_recently_modified.setMinimumWidth(140)
        self._btn_recently_modified.setToolTip("Show instruments modified in the last 7 days")
        self._btn_recently_modified.clicked.connect(self._on_needs_attention_recently_modified)
        needs_layout.addWidget(self._btn_recently_modified)
        self._btn_failed = QtWidgets.QPushButton("Last cal failed (0)")
        self._btn_failed.setCheckable(True)
        self._btn_failed.setMinimumWidth(115)
        self._btn_failed.setToolTip("Show instruments whose most recent calibration failed")
        self._btn_failed.clicked.connect(self._on_needs_attention_failed)
        needs_layout.addWidget(self._btn_failed)
        needs_layout.addSpacing(12)
        self._btn_clear_attention = QtWidgets.QPushButton("Clear")
        self._btn_clear_attention.setMinimumWidth(55)
        self._btn_clear_attention.setToolTip("Clear Needs Attention filter")
        self._btn_clear_attention.clicked.connect(self._on_needs_attention_clear)
        needs_layout.addWidget(self._btn_clear_attention)
        needs_layout.addStretch(1)
        layout.addWidget(self._needs_attention_container)

        # ------------------------------------------------------------------
        # Filters row with better layout
        # ------------------------------------------------------------------
        filters_container = QtWidgets.QWidget()
        filters_layout = QtWidgets.QHBoxLayout(filters_container)
        filters_layout.setContentsMargins(5, 5, 5, 5)
        filters_layout.setSpacing(10)

        search_label = QtWidgets.QLabel("Search:")
        search_label.setToolTip("Search instruments")
        filters_layout.addWidget(search_label)
        self.search_edit = QtWidgets.QLineEdit()
        self.search_edit.setPlaceholderText("Search by ID, location, destination, or type...")
        self.search_edit.setToolTip("Type to search instruments. Press Ctrl+F to focus.")
        self.search_edit.setClearButtonEnabled(True)
        filters_layout.addWidget(self.search_edit, 3)

        # Add Ctrl+F shortcut for search
        search_shortcut = QtWidgets.QShortcut(QtGui.QKeySequence("Ctrl+F"), self)
        search_shortcut.activated.connect(lambda: self.search_edit.setFocus())
        
        # Update highlight delegate when search text changes
        self.search_edit.textChanged.connect(self._update_search_highlight)

        filters_layout.addSpacing(10)

        self.status_filter_combo = QtWidgets.QComboBox()
        self.status_filter_combo.addItem("All Statuses", "")
        self.status_filter_combo.addItem("ACTIVE", "ACTIVE")
        self.status_filter_combo.addItem("RETIRED", "RETIRED")
        self.status_filter_combo.addItem("OUT_FOR_CAL", "OUT_FOR_CAL")
        self.status_filter_combo.setToolTip("Filter by instrument status")
        filters_layout.addWidget(QtWidgets.QLabel("Status:"))
        filters_layout.addWidget(self.status_filter_combo)

        self.type_filter_combo = QtWidgets.QComboBox()
        self.type_filter_combo.addItem("All Types", "")
        for t in self.repo.list_instrument_types():
            self.type_filter_combo.addItem(t["name"], t["name"])
        self.type_filter_combo.setToolTip("Filter by instrument type")
        filters_layout.addWidget(QtWidgets.QLabel("Type:"))
        filters_layout.addWidget(self.type_filter_combo)

        self.due_filter_combo = QtWidgets.QComboBox()
        self.due_filter_combo.addItem("All")
        self.due_filter_combo.addItem("Overdue")
        self.due_filter_combo.addItem("Due in 30 days")
        self.due_filter_combo.setToolTip("Filter by calibration due date")
        filters_layout.addWidget(QtWidgets.QLabel("Due:"))
        filters_layout.addWidget(self.due_filter_combo)

        # Clear filters button
        clear_filters_btn = QtWidgets.QPushButton("Clear")
        clear_filters_btn.setToolTip("Clear all filters")
        clear_filters_btn.setMaximumWidth(60)
        clear_filters_btn.clicked.connect(self._clear_filters)
        filters_layout.addWidget(clear_filters_btn)

        # Show archived instruments (soft-deleted)
        self.show_archived_check = QtWidgets.QCheckBox("Show archived")
        self.show_archived_check.setToolTip("Include archived (soft-deleted) instruments in the list")
        self.show_archived_check.toggled.connect(self.load_instruments)
        filters_layout.addWidget(self.show_archived_check)

        layout.addWidget(filters_container)

        # ------------------------------------------------------------------
        # Table + models
        # ------------------------------------------------------------------
        self.table = QtWidgets.QTableView()
        self.model = InstrumentTableModel([])

        self.proxy = InstrumentFilterProxyModel(self)
        self.proxy.setSourceModel(self.model)

        self.table.setModel(self.proxy)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setSortingEnabled(True)
        self.table.doubleClicked.connect(self.on_table_double_clicked)
        self.table.setAlternatingRowColors(True)
        self.table.setShowGrid(False)
        self.table.verticalHeader().setVisible(False)
        
        # Enable context menu
        self.table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._show_table_context_menu)

        header = self.table.horizontalHeader()
        header.setHighlightSections(False)
        header.setDefaultAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
        header.setSectionResizeMode(QtWidgets.QHeaderView.Interactive)
        self.table.setWordWrap(False)
        
        # Better row height
        self.table.verticalHeader().setDefaultSectionSize(24)
        
        # Enable keyboard navigation
        self.table.setTabKeyNavigation(True)
        
        # Minimum column widths and persistence (must be after header setup)
        MIN_COLUMN_WIDTHS = (80, 120, 70, 120, 90, 90, 70, 90, 120, 72)  # ID, Location, Type, ..., Flag
        for i, min_w in enumerate(MIN_COLUMN_WIDTHS):
            if i < header.count():
                header.setMinimumSectionSize(min_w)
        self._restore_column_widths()
        header.sectionResized.connect(self._save_column_widths)
        
        # Search highlighting delegate
        self.highlight_delegate = HighlightDelegate("", self.table)
        self.table.setItemDelegate(self.highlight_delegate)

        layout.addWidget(self.table)
        
        # Statistics panel
        self.stats_widget = self._create_statistics_widget()
        layout.addWidget(self.stats_widget)
        
        self.setCentralWidget(central)

        # Filter signal wiring
        self.search_edit.textChanged.connect(self._on_filters_changed)
        self.status_filter_combo.currentIndexChanged.connect(self._on_filters_changed)
        self.type_filter_combo.currentIndexChanged.connect(self._on_filters_changed)
        self.due_filter_combo.currentIndexChanged.connect(self._on_filters_changed)
        
        # Update statistics when filters change
        self.proxy.rowsInserted.connect(lambda: self._update_statistics())
        self.proxy.rowsRemoved.connect(lambda: self._update_statistics())
        self.proxy.modelReset.connect(lambda: self._update_statistics())

        # Status bar with better info
        self.statusBar().showMessage("Ready - Select an instrument to get started")

        # Refresh hint: clickable "Refresh" link (helps with stale data / multi-user awareness)
        self._refresh_hint = QtWidgets.QLabel('<a href="#" style="color: inherit;">Refresh</a>')
        self._refresh_hint.setOpenExternalLinks(False)
        self._refresh_hint.linkActivated.connect(lambda _: self._on_refresh_clicked())
        self._refresh_hint.setToolTip("Reload data from database (data may have changed)")
        self.statusBar().addPermanentWidget(self._refresh_hint)

        # Update status bar when selection changes
        self.table.selectionModel().selectionChanged.connect(self._update_status_bar)

    def _on_refresh_clicked(self):
        """Refresh data (triggered by status bar link)."""
        self.load_instruments()
        self.statusBar().showMessage("Data refreshed", 2000)

    def _update_window_title(self):
        """Set window title to show app version (from VERSION / update checker)."""
        try:
            from update_checker import get_current_version
            ver = get_current_version()
            if ver:
                self.setWindowTitle(f"Calibration Tracker - v{ver}")
                return
        except Exception:
            pass
        self.setWindowTitle("Calibration Tracker")

    def load_instruments(self):
        include_archived = getattr(
            self, "show_archived_check", None
        ) and self.show_archived_check.isChecked()
        instruments = self.repo.list_instruments(include_archived=include_archived)
        self.model.set_instruments(instruments)
        count = len(instruments)

        self._update_window_title()
        self.statusBar().showMessage(f"Loaded {count} instrument(s)", 3000)
        self._refresh_hint.setText('<a href="#" style="color: inherit;">Refreshed</a>')
        QtCore.QTimer.singleShot(3000, lambda: self._refresh_hint.setText('<a href="#" style="color: inherit;">Refresh</a>'))
        self._update_statistics()
        self._update_needs_attention_counts()
        if hasattr(self, "_update_needs_attention_button_states"):
            self._update_needs_attention_button_states()
    
    def _update_needs_attention_counts(self):
        """Refresh Overdue / Due soon / Recently modified counts."""
        include_archived = getattr(
            self, "show_archived_check", None
        ) and self.show_archived_check.isChecked()
        try:
            n_overdue = len(self.repo.get_overdue_instruments(include_archived=include_archived))
            n_due_soon = len(self.repo.get_due_soon_instruments(30, include_archived=include_archived))
            n_recent = len(self.repo.get_recently_modified_instruments(7, include_archived=include_archived))
        except Exception:
            n_overdue = n_due_soon = n_recent = 0
        n_failed = sum(
            1 for i in self.model.instruments
            if (i.get("last_cal_result") or "").strip().upper() == "FAIL"
        )
        self._btn_overdue.setText(f"Overdue ({n_overdue})")
        self._btn_due_soon.setText(f"Due soon ({n_due_soon})")
        self._btn_recently_modified.setText(f"Recently modified ({n_recent})")
        self._btn_failed.setText(f"Last cal failed ({n_failed})")

    def _on_needs_attention_overdue(self):
        self.due_filter_combo.setCurrentText("Overdue")
        self.proxy.set_recently_modified_days(0)
        self.proxy.set_failed_only(False)
        self._on_filters_changed()
        self._update_needs_attention_button_states()
        self.statusBar().showMessage("Showing overdue instruments", 2000)

    def _on_needs_attention_due_soon(self):
        self.due_filter_combo.setCurrentText("Due in 30 days")
        self.proxy.set_recently_modified_days(0)
        self.proxy.set_failed_only(False)
        self._on_filters_changed()
        self._update_needs_attention_button_states()
        self.statusBar().showMessage("Showing instruments due within 30 days", 2000)

    def _on_needs_attention_recently_modified(self):
        self.due_filter_combo.setCurrentIndex(0)  # All
        self.proxy.set_due_filter("All")
        self.proxy.set_recently_modified_days(7)
        self.proxy.set_failed_only(False)
        self._on_filters_changed()
        self._update_needs_attention_button_states()
        self.statusBar().showMessage("Showing instruments modified in last 7 days", 2000)

    def _on_needs_attention_failed(self):
        self.due_filter_combo.setCurrentIndex(0)
        self.proxy.set_due_filter("All")
        self.proxy.set_recently_modified_days(0)
        self.proxy.set_failed_only(True)
        self._on_filters_changed()
        self._update_needs_attention_button_states()
        self.statusBar().showMessage("Showing instruments whose last calibration failed", 2000)

    def _on_needs_attention_clear(self):
        self.due_filter_combo.setCurrentIndex(0)
        self.proxy.set_due_filter("All")
        self.proxy.set_recently_modified_days(0)
        self.proxy.set_failed_only(False)
        self._on_filters_changed()
        self._update_needs_attention_button_states()
        self.statusBar().showMessage("Needs Attention filter cleared", 2000)

    def _update_needs_attention_button_states(self):
        """Highlight active Needs Attention filter button."""
        self._btn_overdue.setChecked(self.due_filter_combo.currentText() == "Overdue")
        self._btn_due_soon.setChecked(self.due_filter_combo.currentText() == "Due in 30 days")
        self._btn_recently_modified.setChecked(self.proxy.recently_modified_days > 0)
        self._btn_failed.setChecked(getattr(self.proxy, "failed_only", False))

    def _clear_filters(self):
        """Clear all filters and reset search."""
        self.search_edit.clear()
        self.status_filter_combo.setCurrentIndex(0)
        self.type_filter_combo.setCurrentIndex(0)
        self.due_filter_combo.setCurrentIndex(0)
        self.proxy.set_recently_modified_days(0)
        self.proxy.set_failed_only(False)
        self._on_filters_changed()
        self._update_needs_attention_button_states()
        self.statusBar().showMessage("Filters cleared", 2000)
    
    def _update_status_bar(self):
        """Update status bar with information about selected instrument."""
        inst_id = self._selected_instrument_id()
        if inst_id:
            inst = self.repo.get_instrument(inst_id)
            if inst:
                tag = inst.tag_number or ""
                location = inst.location or ""
                next_due = inst.next_due_date or ""
                status_msg = f"Selected: {tag}"
                if location:
                    status_msg += f" | Location: {location}"
                if next_due:
                    status_msg += f" | Next due: {next_due}"
                self.statusBar().showMessage(status_msg)
            else:
                self.statusBar().showMessage("Ready")
        else:
            self.statusBar().showMessage("Ready - Select an instrument to get started")
    
    def _show_table_context_menu(self, position):
        """Show context menu for table."""
        menu = QtWidgets.QMenu(self)
        
        inst_id = self._selected_instrument_id()
        ids = self._selected_instrument_ids()
        if inst_id:
            menu.addAction(self.act_edit)
            menu.addAction(self.act_view_info)
            menu.addSeparator()
            menu.addAction(self.act_cal)
            menu.addAction(self.act_hist)
            if len(ids) > 1:
                menu.addSeparator()
                menu.addAction(self.act_batch_update)
                menu.addAction(self.act_batch_assign_type)
            menu.addSeparator()
            menu.addAction(self.act_delete)
        else:
            menu.addAction(self.act_new)
        
        menu.exec_(self.table.viewport().mapToGlobal(position))
    
    def on_view_info(self):
        """View detailed information about selected instrument."""
        inst_id = self._selected_instrument_id()
        if not inst_id:
            QtWidgets.QMessageBox.information(
                self,
                "No selection",
                "Please select an instrument to view details.",
            )
            return
        dlg = InstrumentInfoDialog(self.repo, inst_id, parent=self)
        dlg.exec_()

    def _selected_instrument_id(self):
        idx = self.table.currentIndex()
        if not idx.isValid():
            return None
        src_idx = self.proxy.mapToSource(idx)
        row = src_idx.row()
        return self.model.get_instrument_id(row)

    def _selected_instrument_ids(self):
        """Return list of instrument IDs for all currently selected rows (multi-select)."""
        ids = []
        for idx in self.table.selectionModel().selectedRows():
            if not idx.isValid():
                continue
            src_idx = self.proxy.mapToSource(idx)
            row = src_idx.row()
            iid = self.model.get_instrument_id(row)
            if iid and iid not in ids:
                ids.append(iid)
        return ids

    def on_table_double_clicked(self, index: QtCore.QModelIndex):
        if not index.isValid():
            return
        src_idx = self.proxy.mapToSource(index)
        row = src_idx.row()
        inst_id = self.model.get_instrument_id(row)
        if not inst_id:
            return
        dlg = InstrumentInfoDialog(self.repo, inst_id, parent=self)
        dlg.exec_()


    def on_new(self):
        dlg = InstrumentDialog(self.repo, parent=self)
        if dlg.exec_() == QtWidgets.QDialog.Accepted:
            data = dlg.get_data()
            if data:
                try:
                    QtWidgets.QApplication.setOverrideCursor(QtCore.Qt.WaitCursor)
                    inst_id = instrument_service.add_instrument(self.repo, data)
                    self.load_instruments()
                    self.statusBar().showMessage("New instrument created successfully", 3000)
                except Exception as e:
                    QtWidgets.QMessageBox.critical(
                        self,
                        "Creation failed",
                        f"Failed to create instrument:\n{str(e)}",
                    )
                finally:
                    QtWidgets.QApplication.restoreOverrideCursor()

    def on_batch_update(self):
        ids = self._selected_instrument_ids()
        if not ids:
            QtWidgets.QMessageBox.information(
                self,
                "No selection",
                "Please select one or more instruments (Ctrl+click or Shift+click for multi-select).",
            )
            return
        dlg = BatchUpdateDialog(len(ids), parent=self)
        if dlg.exec_() != QtWidgets.QDialog.Accepted:
            return
        updates, reason = dlg.get_update()
        if not updates:
            return
        try:
            self.repo.batch_update_instruments(ids, updates, reason=reason)
            self.load_instruments()
            self.statusBar().showMessage(f"Updated {len(ids)} instrument(s)", 3000)
        except Exception as e:
            QtWidgets.QMessageBox.critical(
                self,
                "Batch update failed",
                str(e),
            )

    def on_batch_assign_type(self):
        ids = self._selected_instrument_ids()
        if not ids:
            QtWidgets.QMessageBox.information(
                self,
                "No selection",
                "Please select one or more instruments (Ctrl+click or Shift+click for multi-select).",
            )
            return
        dlg = BatchAssignInstrumentTypeDialog(self.repo, len(ids), parent=self)
        if dlg.exec_() != QtWidgets.QDialog.Accepted:
            return
        type_id, reason = dlg.get_result()
        try:
            self.repo.batch_update_instruments(
                ids, {"instrument_type_id": type_id}, reason=reason
            )
            self.load_instruments()
            self.statusBar().showMessage(
                f"Assigned instrument type to {len(ids)} instrument(s)", 3000
            )
        except Exception as e:
            QtWidgets.QMessageBox.critical(
                self,
                "Batch assign instrument type failed",
                str(e),
            )

    def on_edit(self):
        inst_id = self._selected_instrument_id()
        if not inst_id:
            QtWidgets.QMessageBox.information(
                self,
                "No selection",
                "Please select an instrument to edit.",
            )
            return
        inst = self.repo.get_instrument(inst_id)
        dlg = InstrumentDialog(self.repo, inst, parent=self)
        if dlg.exec_() == QtWidgets.QDialog.Accepted:
            data = dlg.get_data()
            if data:
                try:
                    QtWidgets.QApplication.setOverrideCursor(QtCore.Qt.WaitCursor)
                    instrument_service.update_instrument(self.repo, inst_id, data)
                    self.load_instruments()
                    self.statusBar().showMessage("Instrument updated successfully", 3000)
                except StaleDataError as e:
                    QtWidgets.QMessageBox.warning(
                        self,
                        "Update failed",
                        str(e) + "\n\nRefresh the list and try again.",
                    )
                    self.load_instruments()
                except Exception as e:
                    QtWidgets.QMessageBox.critical(
                        self,
                        "Update failed",
                        f"Failed to update instrument:\n{str(e)}",
                    )
                finally:
                    QtWidgets.QApplication.restoreOverrideCursor()

    def on_delete(self):
        inst_id = self._selected_instrument_id()
        if not inst_id:
            QtWidgets.QMessageBox.information(
                self,
                "No selection",
                "Please select an instrument to delete.",
            )
            return

        inst = self.repo.get_instrument(inst_id)
        tag = inst.tag_number if inst else str(inst_id)
        location = inst.location or "" if inst else ""

        msg = (
            f"Instrument: {tag}\n"
            f"Location: {location or '(none)'}\n\n"
            "Choose an action:\n\n"
            "• Archive: Hide from list but keep all history (recommended)\n"
            "• Delete permanently: Remove completely (cannot be undone)"
        )

        box = QtWidgets.QMessageBox(self)
        box.setIcon(QtWidgets.QMessageBox.Warning)
        box.setWindowTitle("Delete instrument")
        box.setText(msg)
        archive_btn = box.addButton("Archive", QtWidgets.QMessageBox.ActionRole)
        delete_btn = box.addButton("Delete permanently", QtWidgets.QMessageBox.DestructiveRole)
        cancel_btn = box.addButton("Cancel", QtWidgets.QMessageBox.RejectRole)
        box.setDefaultButton(archive_btn)
        delete_btn.setStyleSheet("QPushButton { background-color: #d32f2f; color: white; }")
        box.exec_()

        clicked = box.clickedButton()
        if clicked == cancel_btn:
            return
        if clicked == archive_btn:
            try:
                self.repo.archive_instrument(inst_id, reason="Archived from delete dialog")
                self.load_instruments()
                self.statusBar().showMessage(f"Archived instrument '{tag}'", 3000)
            except Exception as e:
                QtWidgets.QMessageBox.critical(
                    self,
                    "Archive failed",
                    f"Failed to archive instrument:\n{str(e)}\n\nPlease try again or contact support if the problem persists.",
                )
            return

        # Delete permanently
        reason, ok = QtWidgets.QInputDialog.getText(
            self,
            "Reason for deletion",
            "Reason (optional, for audit log):",
            QtWidgets.QLineEdit.Normal,
            "",
        )
        if not ok:
            return
        reason = reason.strip() or None

        try:
            self.repo.delete_instrument(inst_id, reason=reason)
            self.load_instruments()
            self.statusBar().showMessage(f"Deleted instrument '{tag}'", 3000)
        except Exception as e:
            QtWidgets.QMessageBox.critical(
                self,
                "Delete failed",
                f"Failed to delete instrument:\n{str(e)}\n\nPlease try again or contact support if the problem persists.",
            )

    def on_mark_calibrated(self):
        inst_id = self._selected_instrument_id()
        if not inst_id:
            QtWidgets.QMessageBox.information(
                self,
                "No selection",
                "Please select an instrument to mark as calibrated.",
            )
            return

        # Prefill dialog with existing last_cal_date if present
        inst = self.repo.get_instrument(inst_id)
        initial_qdate = None
        if inst and inst.last_cal_date:
            try:
                d = datetime.strptime(inst.last_cal_date, "%Y-%m-%d").date()
                initial_qdate = QtCore.QDate(d.year, d.month, d.day)
            except Exception:
                pass

        dlg = CalDateDialog(self, initial_date=initial_qdate)
        if dlg.exec_() != QtWidgets.QDialog.Accepted:
            return

        picked_date = dlg.get_date()
        try:
            # Update DB: last_cal_date = picked_date, next_due_date = picked_date + 1 year
            self.repo.mark_calibrated_on(inst_id, picked_date)
            self.load_instruments()
            tag = inst.tag_number if inst else "instrument"
            self.statusBar().showMessage(f"{tag} marked as calibrated on {picked_date.isoformat()}", 3000)
        except Exception as e:
            QtWidgets.QMessageBox.critical(
                self,
                "Update failed",
                f"Failed to mark instrument as calibrated:\n{str(e)}",
            )

    def on_destinations(self):
        dlg = DestinationsDialog(self.repo, parent=self)
        dlg.exec_()
        # Refresh instrument list so new/renamed destinations show up in table
        self.load_instruments()

    def on_personnel(self):
        dlg = PersonnelDialog(self.repo, parent=self)
        dlg.exec_()

    def on_cal_history(self):
        inst_id = self._selected_instrument_id()
        if not inst_id:
            QtWidgets.QMessageBox.information(
                self,
                "No selection",
                "Please select an instrument to view calibration history.",
            )
            return
        dlg = CalibrationHistoryDialog(self.repo, inst_id, parent=self)
        dlg.exec_()
        # Refresh in case calibrations were added/modified
        self.load_instruments()

    def on_templates(self):
        dlg = TemplatesDialog(self.repo, parent=self)
        dlg.exec_()

    def on_send_reminders(self):
        self.act_reminders.setEnabled(False)
        try:
            count = send_due_reminders_via_lan(self.repo)
            if count:
                QtWidgets.QMessageBox.information(
                    self,
                    "Reminders sent",
                    f"LAN reminder broadcast sent for {count} instrument(s).",
                )
            else:
                QtWidgets.QMessageBox.information(
                    self,
                    "No reminders",
                    "No instruments due within the reminder window.",
                )
        finally:
            self.act_reminders.setEnabled(True)

    def on_settings(self):
        dlg = SettingsDialog(self.repo, parent=self)
        if dlg.exec_() == QtWidgets.QDialog.Accepted:
            self.statusBar().showMessage("Settings saved", 3000)

    def _on_filters_changed(self):
        self.proxy.set_text_filter(self.search_edit.text())
        self.proxy.set_status_filter(self.status_filter_combo.currentData() or "")
        self.proxy.set_type_filter(self.type_filter_combo.currentData() or "")
        self.proxy.set_due_filter(self.due_filter_combo.currentText())

        self._update_window_title()
        self._update_statistics()
        self._update_needs_attention_button_states()

    def on_export_csv(self):
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Export current view to CSV",
            "",
            "CSV files (*.csv);;All files (*)",
        )
        if not path:
            return

        if Path(path).exists():
            reply = QtWidgets.QMessageBox.question(
                self,
                "File exists",
                f"The file already exists:\n{path}\n\nOverwrite?",
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No | QtWidgets.QMessageBox.Cancel,
                QtWidgets.QMessageBox.No,
            )
            if reply != QtWidgets.QMessageBox.Yes:
                return

        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)

                # headers
                headers = [
                    self.proxy.headerData(c, QtCore.Qt.Horizontal, QtCore.Qt.DisplayRole)
                    for c in range(self.proxy.columnCount())
                ]
                writer.writerow(headers)

                # rows (filtered/sorted view)
                for row in range(self.proxy.rowCount()):
                    row_vals = []
                    for col in range(self.proxy.columnCount()):
                        idx = self.proxy.index(row, col)
                        val = self.proxy.data(idx, QtCore.Qt.DisplayRole)
                        row_vals.append("" if val is None else str(val))
                    writer.writerow(row_vals)

            QtWidgets.QMessageBox.information(
                self,
                "Export complete",
                f"Exported {self.proxy.rowCount()} row(s) to:\n{path}",
            )
        except Exception as e:
            QtWidgets.QMessageBox.critical(
                self,
                "Export failed",
                str(e),
            )

    def on_export_excel(self):
        """Export current (filtered) instrument view to XLSX using openpyxl."""
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Export current view to Excel",
            "",
            "Excel files (*.xlsx);;All files (*)",
        )
        if not path:
            return
        if not path.endswith(".xlsx"):
            path = path + ".xlsx"

        if Path(path).exists():
            reply = QtWidgets.QMessageBox.question(
                self,
                "File exists",
                f"The file already exists:\n{path}\n\nOverwrite?",
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No | QtWidgets.QMessageBox.Cancel,
                QtWidgets.QMessageBox.No,
            )
            if reply != QtWidgets.QMessageBox.Yes:
                return

        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            if ws is None:
                ws = wb.create_sheet("Instruments", 0)
            ws.title = "Instruments"
            # Headers
            headers = [
                self.proxy.headerData(c, QtCore.Qt.Horizontal, QtCore.Qt.DisplayRole)
                for c in range(self.proxy.columnCount())
            ]
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h or "")
            # Rows (filtered/sorted view)
            for row in range(self.proxy.rowCount()):
                for col in range(self.proxy.columnCount()):
                    idx = self.proxy.index(row, col)
                    val = self.proxy.data(idx, QtCore.Qt.DisplayRole)
                    ws.cell(row=row + 2, column=col + 1, value="" if val is None else str(val))
            wb.save(path)
            QtWidgets.QMessageBox.information(
                self,
                "Export complete",
                f"Exported {self.proxy.rowCount()} row(s) to:\n{path}",
            )
        except Exception as e:
            QtWidgets.QMessageBox.critical(
                self,
                "Export failed",
                str(e),
            )

    def _on_export_preset(self, preset_name: str):
        """Export all calibrations to PDF in a preset subfolder (e.g. Internal_Review_2025-01-29)."""
        base_dir = QtWidgets.QFileDialog.getExistingDirectory(
            self,
            f"Select directory for '{preset_name}' export",
            "",
            QtWidgets.QFileDialog.ShowDirsOnly | QtWidgets.QFileDialog.DontResolveSymlinks,
        )
        if not base_dir:
            return
        subfolder = f"{preset_name}_{date.today().isoformat()}"
        import os
        target_dir = os.path.join(base_dir, subfolder)
        reply = QtWidgets.QMessageBox.question(
            self,
            "Export preset",
            f"Export all calibration records to PDF in:\n{target_dir}\n\nContinue?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No,
        )
        if reply != QtWidgets.QMessageBox.Yes:
            return
        self._run_pdf_export_async(target_dir)
    
    def on_export_all_calibrations(self):
        """Export all calibration records to PDF files organized by instrument type."""
        base_dir = QtWidgets.QFileDialog.getExistingDirectory(
            self,
            "Select directory for calibration exports",
            "",
            QtWidgets.QFileDialog.ShowDirsOnly | QtWidgets.QFileDialog.DontResolveSymlinks,
        )
        if not base_dir:
            return
        reply = QtWidgets.QMessageBox.question(
            self,
            "Export all calibrations",
            f"This will export all calibration records to PDF files in:\n{base_dir}\n\n"
            f"Files will be organized by instrument type.\n\n"
            f"Continue?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No,
        )
        if reply != QtWidgets.QMessageBox.Yes:
            return
        self._run_pdf_export_async(base_dir)

    def _run_pdf_export_async(self, target_dir: str | Path):
        """Run PDF export in background thread with progress dialog and cancel support."""
        target_dir = Path(target_dir)
        db_path = get_effective_db_path()
        self._export_worker = ExportWorker(db_path, target_dir)
        self._export_progress = QtWidgets.QProgressDialog(
            "Exporting calibrations...",
            "Cancel",
            0,
            0,
            self,
        )
        self._export_progress.setWindowModality(QtCore.Qt.WindowModal)
        self._export_progress.setMinimumDuration(0)
        self._export_progress.canceled.connect(self._on_export_cancel)
        self._export_worker.progress.connect(self._on_export_progress)
        self._export_worker.finished.connect(self._on_export_finished)
        self._export_worker.error.connect(self._on_export_error)
        self._export_progress.show()
        self._export_worker.start()

    def _on_export_progress(self, current: int, total: int):
        if self._export_progress.maximum() == 0 and total > 0:
            self._export_progress.setMaximum(total)
        self._export_progress.setValue(current)
        self._export_progress.setLabelText(f"Exporting {current}/{total}...")

    def _on_export_cancel(self):
        self._export_worker.cancel()
        self._export_progress.setLabelText("Cancelling...")

    def _on_export_finished(self, result: dict):
        self._export_progress.close()
        self._export_progress = None
        self._export_worker = None
        cancelled = result.get("cancelled", False)
        if cancelled:
            msg = (
                f"Export cancelled.\n\n"
                f"Partially exported: {result['success_count']} calibration(s)\n"
                f"Errors: {result['error_count']}"
            )
        else:
            msg = (
                f"Export complete!\n\n"
                f"Successfully exported: {result['success_count']} calibration(s)\n"
                f"Attachments exported: {result.get('attachment_count', 0)}\n"
                f"Errors: {result['error_count']}"
            )
        if result.get("errors"):
            error_details = "\n".join(result["errors"][:10])
            if len(result["errors"]) > 10:
                error_details += f"\n... and {len(result['errors']) - 10} more errors"
            msg += f"\n\nErrors:\n{error_details}"
        if result["error_count"] > 0:
            QtWidgets.QMessageBox.warning(self, "Export complete with errors", msg)
        else:
            QtWidgets.QMessageBox.information(
                self, "Export cancelled" if cancelled else "Export complete", msg
            )

    def _on_export_error(self, err_msg: str):
        if self._export_progress:
            self._export_progress.close()
            self._export_progress = None
        self._export_worker = None
        QtWidgets.QMessageBox.critical(
            self, "Export failed", f"Error exporting calibrations:\n{err_msg}"
        )
    
    def _create_statistics_widget(self):
        """Create a statistics panel showing key metrics."""
        stats = QtWidgets.QWidget()
        stats_layout = QtWidgets.QHBoxLayout(stats)
        stats_layout.setContentsMargins(8, 4, 8, 4)
        stats_layout.setSpacing(15)
        
        self.total_label = QtWidgets.QLabel("Total: 0")
        self.active_label = QtWidgets.QLabel("Active: 0")
        self.overdue_label = QtWidgets.QLabel("Overdue: 0")
        self.due_soon_label = QtWidgets.QLabel("Due in 30 days: 0")
        self.failed_label = QtWidgets.QLabel("Last cal failed: 0")
        self.failed_label.setToolTip("Instruments whose most recent calibration overall result was Fail (one per instrument)")
        
        # Style the labels (base styling)
        for label in [self.total_label, self.active_label, self.overdue_label, self.due_soon_label, self.failed_label]:
            label.setStyleSheet("padding: 4px 8px; border-radius: 3px;")
        self._apply_statistics_colors()
        
        stats_layout.addWidget(self.total_label)
        stats_layout.addWidget(self.active_label)
        stats_layout.addWidget(self.overdue_label)
        stats_layout.addWidget(self.due_soon_label)
        stats_layout.addWidget(self.failed_label)
        stats_layout.addStretch()
        
        return stats
    
    def _update_statistics(self):
        """Update statistics panel with current data. Failed = instruments whose most recent calibration overall result is Fail (one per instrument)."""
        instruments = self.repo.list_instruments()
        total = len(instruments)
        
        active = sum(1 for inst in instruments if inst.get("status") == "ACTIVE")
        
        overdue = 0
        due_soon = 0
        today = date.today()
        
        for inst in instruments:
            next_due = inst.get("next_due_date")
            if next_due:
                try:
                    due_date = datetime.strptime(next_due, "%Y-%m-%d").date()
                    days_left = (due_date - today).days
                    if days_left < 0:
                        overdue += 1
                    elif days_left <= 30:
                        due_soon += 1
                except Exception:
                    pass

        failed = sum(1 for inst in instruments if (str(inst.get("last_cal_result") or "").strip().upper() == "FAIL"))

        self.total_label.setText(f"Total: {total}")
        self.active_label.setText(f"Active: {active}")
        self.overdue_label.setText(f"Overdue: {overdue}")
        self.due_soon_label.setText(f"Due in 30 days: {due_soon}")
        self.failed_label.setText(f"Last cal failed: {failed}")

    def _apply_statistics_colors(self):
        """Apply theme-derived colors to statistics labels."""
        overdue_color = "#FF9800"  # Orange
        self.overdue_label.setStyleSheet(
            f"padding: 4px 8px; border-radius: 3px; color: {overdue_color}; font-weight: bold;"
        )
        due_soon_color = "#FFEB3B"  # Yellow for upcoming
        self.due_soon_label.setStyleSheet(
            f"padding: 4px 8px; border-radius: 3px; color: {due_soon_color};"
        )
        failed_color = "#FF4444"  # Bright red for cal fails
        self.failed_label.setStyleSheet(
            f"padding: 4px 8px; border-radius: 3px; color: {failed_color}; font-weight: bold;"
        )
    
    def _update_search_highlight(self):
        """Update the search highlight delegate with current search text."""
        search_text = self.search_edit.text()
        self.highlight_delegate.set_search_text(search_text)
        # Trigger repaint of visible cells
        self.table.viewport().update()
    
    def _save_column_widths(self):
        """Save column widths to settings."""
        settings = QtCore.QSettings("CalibrationTracker", "ColumnWidths")
        header = self.table.horizontalHeader()
        for i in range(header.count()):
            width = header.sectionSize(i)
            settings.setValue(f"column_{i}", width)
    
    def _restore_column_widths(self):
        """Restore column widths from settings, enforcing minimums."""
        settings = QtCore.QSettings("CalibrationTracker", "ColumnWidths")
        header = self.table.horizontalHeader()
        MIN_COLUMN_WIDTHS = (80, 120, 70, 120, 90, 90, 70, 90, 120)
        for i in range(header.count()):
            width = settings.value(f"column_{i}", None)
            min_w = MIN_COLUMN_WIDTHS[i] if i < len(MIN_COLUMN_WIDTHS) else 60
            if width is not None:
                try:
                    header.resizeSection(i, max(int(width), min_w))
                except (ValueError, TypeError):
                    header.resizeSection(i, min_w)
            elif i < len(MIN_COLUMN_WIDTHS):
                header.resizeSection(i, min_w)

    def _populate_theme_menu(self):
        """Rebuild theme submenu from get_all_themes() (includes custom themes)."""
        for act in self._theme_action_group.actions()[:]:
            self._theme_action_group.removeAction(act)
        self._theme_sub.clear()
        current_theme = get_saved_theme()
        for theme_name in sorted(get_all_themes().keys(), key=lambda n: n.lower()):
            act = self._theme_sub.addAction(theme_name)
            act.setCheckable(True)
            act.setChecked(theme_name == current_theme)
            act.triggered.connect(lambda checked, n=theme_name: self._on_theme_selected(n))
            self._theme_action_group.addAction(act)
        self._theme_sub.addSeparator()
        customize_act = self._theme_sub.addAction("Customize themes...")
        customize_act.triggered.connect(self._on_customize_themes)

    def _on_customize_themes(self):
        dlg = ThemeEditorDialog(self)
        dlg.exec_()

    def _on_theme_selected(self, theme_name: str):
        """Apply the selected theme and remember it for next run."""
        set_saved_theme(theme_name)
        app = QtWidgets.QApplication.instance()
        if app:
            apply_global_style(app, theme_name)
        for act in self._theme_action_group.actions():
            act.setChecked(act.text() == theme_name)
        if hasattr(self, "_apply_statistics_colors"):
            self._apply_statistics_colors()

    def _on_text_size_selected(self, pt: int):
        """Apply the selected text size and remember it for next run."""
        set_saved_font_size(pt)
        app = QtWidgets.QApplication.instance()
        if app:
            apply_global_style(app)
        for act in self._text_size_action_group.actions():
            act.setChecked(act.data() == pt)
        if hasattr(self, "_apply_statistics_colors"):
            self._apply_statistics_colors()

    def _on_refresh_database(self):
        """Try to reconnect to the server database; refresh UI if successful."""
        # Use shorter timeout so UI doesn't hang long if server is unreachable (10s instead of 30s).
        QtWidgets.QApplication.setOverrideCursor(QtCore.Qt.WaitCursor)
        try:
            try:
                new_conn = get_connection(DB_PATH, timeout=10.0)
            except Exception as e:
                QtWidgets.QMessageBox.warning(
                    self,
                    "Refresh database",
                    f"Could not connect to the server database:\n{e}\n\nStill using current database.",
                )
                return
            try:
                # Initialize schema on the new connection (same as startup).
                from database import initialize_db
                new_conn = initialize_db(new_conn, DB_PATH)
            except Exception as e:
                try:
                    new_conn.close()
                except Exception:
                    pass
                QtWidgets.QMessageBox.warning(
                    self,
                    "Refresh database",
                    f"Database initialization failed:\n{e}\n\nStill using current database.",
                )
                return
            # Close old connection first so we don't hold two connections.
            old_conn = self.repo.conn
            try:
                old_conn.close()
            except Exception:
                pass
            self.repo = CalibrationRepository(new_conn)
            persist_last_db_path(get_effective_db_path())
            self.load_instruments()
            self._update_window_title()
            self._update_statistics()
            QtWidgets.QMessageBox.information(
                self,
                "Refresh database",
                "Reconnected to the server database. Data has been reloaded.",
            )
        finally:
            QtWidgets.QApplication.restoreOverrideCursor()

    def on_show_shortcuts(self):
        """Show keyboard shortcuts dialog."""
        shortcuts = """
        <h2>Keyboard Shortcuts</h2>
        <table style="width: 100%; border-collapse: collapse;">
        <tr><td style="padding: 5px;"><b>Ctrl+N</b></td><td style="padding: 5px;">Create new instrument</td></tr>
        <tr><td style="padding: 5px;"><b>Ctrl+E</b></td><td style="padding: 5px;">Edit selected instrument</td></tr>
        <tr><td style="padding: 5px;"><b>Delete</b></td><td style="padding: 5px;">Delete selected instrument</td></tr>
        <tr><td style="padding: 5px;"><b>Ctrl+I</b></td><td style="padding: 5px;">View instrument details</td></tr>
        <tr><td style="padding: 5px;"><b>Ctrl+F</b></td><td style="padding: 5px;">Focus search box</td></tr>
        <tr><td style="padding: 5px;"><b>Ctrl+M</b></td><td style="padding: 5px;">Mark instrument as calibrated</td></tr>
        <tr><td style="padding: 5px;"><b>Ctrl+H</b></td><td style="padding: 5px;">View calibration history</td></tr>
        <tr><td style="padding: 5px;"><b>Ctrl+T</b></td><td style="padding: 5px;">Open templates dialog</td></tr>
        <tr><td style="padding: 5px;"><b>Ctrl+,</b></td><td style="padding: 5px;">Open settings</td></tr>
        <tr><td style="padding: 5px;"><b>Ctrl+Shift+C</b></td><td style="padding: 5px;">Export current view to CSV</td></tr>
        <tr><td style="padding: 5px;"><b>Ctrl+Shift+P</b></td><td style="padding: 5px;">Export all calibrations to PDF</td></tr>
        <tr><td style="padding: 5px;"><b>F1</b></td><td style="padding: 5px;">Show this help</td></tr>
        <tr><td style="padding: 5px;"><b>Ctrl+Q</b></td><td style="padding: 5px;">Exit application</td></tr>
        </table>
        <p><i>Tip: Double-click a row to view instrument details</i></p>
        """
        dlg = HelpDialog("Keyboard Shortcuts", shortcuts, self)
        dlg.exec_()
    
    def on_show_about(self):
        """Show about dialog."""
        try:
            from update_checker import get_current_version
            ver = get_current_version() or "—"
        except Exception:
            ver = "—"
        about_text = f"""
        <h2>Calibration Tracker</h2>
        <p><b>Version:</b> {ver}</p>
        <p>A comprehensive application for managing instrument calibrations, tracking schedules, and maintaining compliance records.</p>
        
        <h3>Features</h3>
        <ul>
        <li>Instrument management and tracking</li>
        <li>Calibration record keeping</li>
        <li>Template-based calibration forms</li>
        <li>PDF and CSV export</li>
        <li>Visual indicators for due dates</li>
        <li>LAN reminder notifications</li>
        <li>Complete audit trail</li>
        </ul>
        
        <p><i>Built with PyQt5 and SQLite</i></p>
        <p style="margin-top: 20px;"><small>© 2024 Calibration Tracker</small></p>
        """
        dlg = HelpDialog("About Calibration Tracker", about_text, self)
        dlg.exec_()
